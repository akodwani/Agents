"""Stage B model executor: compute light valuation outputs and export Excel."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("/workspace/Agents/outputs")


def _compute_scenario_metrics(scenario: str, assumptions: Dict[str, float], model_type: str) -> Dict[str, float | str]:
    revenue = assumptions["revenue"]
    growth = assumptions["growth_rate"]
    margin = assumptions["ebitda_margin"]
    tax_rate = assumptions["tax_rate"]
    discount_rate = assumptions["discount_rate"]
    terminal_growth = assumptions["terminal_growth"]

    next_revenue = revenue * (1 + growth)
    ebitda = next_revenue * margin
    nopat = ebitda * (1 - tax_rate)

    if model_type == "3statement-lite":
        value = nopat * 8.0
    else:
        terminal_cashflow = nopat * (1 + terminal_growth)
        denom = max(1e-6, discount_rate - terminal_growth)
        terminal_value = terminal_cashflow / denom
        value = terminal_value / (1 + discount_rate)

    return {
        "scenario": scenario,
        "revenue": round(revenue, 2),
        "next_revenue": round(next_revenue, 2),
        "growth_rate": round(growth, 4),
        "ebitda_margin": round(margin, 4),
        "ebitda": round(ebitda, 2),
        "tax_rate": round(tax_rate, 4),
        "nopat": round(nopat, 2),
        "discount_rate": round(discount_rate, 4),
        "terminal_growth": round(terminal_growth, 4),
        "implied_value": round(value, 2),
        "model_type": model_type,
    }


def _light_format_excel(xlsx_path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 30)
    wb.save(xlsx_path)


def execute_model(assumptions_json: str, model_type: str = "dcf") -> Dict[str, Any]:
    """Compute scenario outputs and write an XLSX file to /outputs/."""
    assumptions = json.loads(assumptions_json)
    rows = [
        _compute_scenario_metrics(scenario_name, scenario_assumptions, model_type)
        for scenario_name, scenario_assumptions in assumptions.items()
    ]

    metrics_df = pd.DataFrame(rows).sort_values("scenario").reset_index(drop=True)

    summary_df = pd.DataFrame(
        {
            "metric": ["avg_implied_value", "max_implied_value", "min_implied_value"],
            "value": [
                round(metrics_df["implied_value"].mean(), 2),
                round(metrics_df["implied_value"].max(), 2),
                round(metrics_df["implied_value"].min(), 2),
            ],
        }
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUTPUT_DIR / f"model_output_{model_type.replace('-', '_')}.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        metrics_df.to_excel(writer, sheet_name="scenario_metrics", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)

    _light_format_excel(xlsx_path)

    return {
        "metrics_df": metrics_df,
        "summary_df": summary_df,
        "xlsx_path": str(xlsx_path),
    }
